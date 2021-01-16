# Importando a biblioteca
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from pandas import DataFrame

# declarando a planilha
arquivo_excel = Workbook()
planilha = arquivo_excel.active
planilha = arquivo_excel.create_sheet("Gastos")

# adicionando valores a correspondente planilha 
planilha['A1'] = 'Categoria'
planilha['B1'] = 'Valor'
planilha['A2'] = "Restaurante"
planilha['B2'] = 45.99

valores = [
    ("Categoria", 'valores'),
    ("Restaurante", 45.99),
    ("Transporte", 208.45),
    ("Viagem", 558.54)
]


fplanilha = pd.DataFrame(planilha)    # printando-a para ver se esta tudo correto

print(fplanilha)